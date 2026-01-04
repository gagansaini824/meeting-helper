import os
import json
import asyncio
import logging
import io
import csv
import base64
import uuid
from datetime import datetime
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, Form, HTTPException, Depends, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from dotenv import load_dotenv

from deepgram import DeepgramClient
from deepgram.core import EventType
from deepgram.extensions.types.sockets import ListenV1MediaMessage
from openai import OpenAI
from pypdf import PdfReader
from docx import Document as DocxDocument
import openpyxl
import anthropic

from vector_store import PineconeVectorStore
from question_detector import detect_question
from auth import clerk_auth, ClerkUser, require_auth, optional_auth
from database import (
    db, get_or_create_user,
    create_session, get_session, get_user_sessions,
    update_session, append_to_session, delete_session, archive_session,
    log_usage, log_audit
)


# User-specific vector stores
_user_vector_stores: dict = {}


def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """Extract text from various file formats"""
    extension = filename.lower().split('.')[-1]

    try:
        # PDF files
        if extension == 'pdf':
            pdf_reader = PdfReader(io.BytesIO(file_bytes))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text.strip()

        # DOCX files
        elif extension == 'docx':
            doc = DocxDocument(io.BytesIO(file_bytes))
            text = "\n".join([para.text for para in doc.paragraphs])
            return text.strip()

        # Markdown files
        elif extension == 'md':
            return file_bytes.decode('utf-8')

        # CSV files
        elif extension == 'csv':
            text = file_bytes.decode('utf-8')
            reader = csv.reader(io.StringIO(text))
            rows = [', '.join(row) for row in reader]
            return '\n'.join(rows)

        # Excel files
        elif extension in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n=== Sheet: {sheet_name} ===\n"
                for row in sheet.iter_rows(values_only=True):
                    text += ', '.join([str(cell) if cell is not None else '' for cell in row]) + '\n'
            return text.strip()

        # Image files - use Claude Vision
        elif extension in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
            return analyze_image_with_claude(file_bytes, extension)

        # Plain text files (txt, py, js, json, etc.)
        else:
            return file_bytes.decode('utf-8')

    except Exception as e:
        raise ValueError(f"Error parsing {extension} file: {str(e)}")


def analyze_image_with_claude(image_bytes: bytes, extension: str) -> str:
    """Analyze image using OpenAI gpt-4o-mini Vision API"""
    try:
        # Convert image to base64
        image_b64 = base64.b64encode(image_bytes).decode('utf-8')

        # Determine media type for data URL
        media_type_map = {
            'png': 'image/png',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'gif': 'image/gif',
            'webp': 'image/webp'
        }
        media_type = media_type_map.get(extension, 'image/jpeg')

        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=2000,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{image_b64}"
                            }
                        },
                        {
                            "type": "text",
                            "text": "Analyze this image in detail. Describe what you see, any text present, diagrams, charts, or other relevant information. Provide a comprehensive description that can be used for semantic search."
                        }
                    ]
                }
            ]
        )

        return response.choices[0].message.content

    except Exception as e:
        return f"[Image file - analysis failed: {str(e)}]"


def get_vector_store_for_user(user_id: str) -> PineconeVectorStore:
    """Get the Pinecone vector store for a user"""
    if user_id not in _user_vector_stores:
        _user_vector_stores[user_id] = PineconeVectorStore()
    return _user_vector_stores[user_id]

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv(override=True)

# Session-aware state management
class SessionState:
    """State for a single session"""
    def __init__(self, session_id: str, user_id: str):
        self.session_id = session_id
        self.user_id = user_id
        self.transcript: list[dict] = []  # [{speaker: int, text: str, timestamp: str, is_final: bool}]
        self.full_transcript: str = ""
        self.detected_questions: list[dict] = []
        self.answers: list[dict] = []
        self.suggestions: list[dict] = []
        self.conversation_summaries: list[dict] = []  # [{timestamp: str, summary: str, transcript_length: int}]
        self.last_analysis_time: float = 0
        self.last_processed_transcript_length: int = 0
        self.last_summary_transcript_length: int = 0  # Track transcript length at last summary
        self.last_save_time: float = 0
        self.dirty: bool = False  # Track if state needs saving

    def add_utterance(self, speaker: int, text: str, is_final: bool = True):
        entry = {
            "speaker": speaker,
            "text": text,
            "timestamp": datetime.now().isoformat(),
            "is_final": is_final
        }
        self.transcript.append(entry)
        if is_final:
            self.full_transcript += f" {text}"
        self.dirty = True

    def add_question(self, question_text: str):
        question = {
            "text": question_text,
            "timestamp": datetime.now().isoformat(),
            "answered": False
        }
        self.detected_questions.append(question)
        self.dirty = True

    def add_answer(self, question: str, answer: str, sources: list = None):
        answer_entry = {
            "question": question,
            "answer": answer,
            "timestamp": datetime.now().isoformat(),
            "sources": sources or []
        }
        self.answers.append(answer_entry)
        # Mark question as answered
        for q in self.detected_questions:
            if q["text"] == question:
                q["answered"] = True
        self.dirty = True

    def get_recent_transcript(self, chars: int = 3000) -> str:
        return self.full_transcript[-chars:] if len(self.full_transcript) > chars else self.full_transcript

    def add_summary(self, summary: str):
        """Add a conversation summary"""
        summary_entry = {
            "timestamp": datetime.now().isoformat(),
            "summary": summary,
            "transcript_length": len(self.full_transcript)
        }
        self.conversation_summaries.append(summary_entry)
        self.last_summary_transcript_length = len(self.full_transcript)
        self.dirty = True
        logger.info(f"Added conversation summary #{len(self.conversation_summaries)} for session {self.session_id}")

    def get_summaries_text(self) -> str:
        """Get all summaries as formatted text for context"""
        if not self.conversation_summaries:
            return ""
        summaries_text = "CONVERSATION SUMMARIES (chronological order):\n"
        for i, s in enumerate(self.conversation_summaries, 1):
            summaries_text += f"\n[Summary {i} - {s['timestamp'][:16]}]:\n{s['summary']}\n"
        return summaries_text

    def needs_summary(self, min_new_chars: int = 2000) -> bool:
        """Check if enough new content exists to warrant a new summary"""
        new_content_length = len(self.full_transcript) - self.last_summary_transcript_length
        return new_content_length >= min_new_chars

    def get_new_transcript_for_detection(self, max_chars: int = 2500, context_overlap: int = 500) -> tuple[str, bool]:
        """
        Get transcript content for question detection with sliding window overlap.

        Args:
            max_chars: Maximum characters to return
            context_overlap: Number of characters to include from already-processed content
                            for context continuity (helps combine fragmented questions)

        Returns:
            Tuple of (transcript_text, has_new_content)
        """
        current_length = len(self.full_transcript)

        # Check if there's any new content since last detection
        if current_length <= self.last_processed_transcript_length:
            return ("", False)

        # Include overlap from previously processed content for context
        # This helps detect questions that span multiple transcript chunks
        context_start = max(0, self.last_processed_transcript_length - context_overlap)
        new_text = self.full_transcript[context_start:]

        # Update the processed length marker
        self.last_processed_transcript_length = current_length

        # Limit to max_chars if needed
        result = new_text[-max_chars:] if len(new_text) > max_chars else new_text
        return (result, True)

    def clear(self):
        self.transcript = []
        self.full_transcript = ""
        self.detected_questions = []
        self.answers = []
        self.suggestions = []
        self.conversation_summaries = []
        self.last_processed_transcript_length = 0
        self.last_summary_transcript_length = 0
        self.dirty = True

    def to_dict(self) -> dict:
        """Convert state to dictionary for saving"""
        return {
            "transcript_entries": self.transcript,
            "full_transcript": self.full_transcript,
            "detected_questions": self.detected_questions,
            "answers": self.answers,
            "conversation_summaries": self.conversation_summaries
        }

    @classmethod
    def from_session_data(cls, session_id: str, user_id: str, data: dict) -> "SessionState":
        """Create SessionState from database session data"""
        state = cls(session_id, user_id)
        state.transcript = data.get("transcript_entries") or []
        state.full_transcript = data.get("full_transcript") or ""
        state.detected_questions = data.get("detected_questions") or []
        state.answers = data.get("answers") or []
        state.conversation_summaries = data.get("conversation_summaries") or []
        # Calculate processed length from existing transcript
        state.last_processed_transcript_length = len(state.full_transcript)
        # Set summary length from last summary if exists
        if state.conversation_summaries:
            state.last_summary_transcript_length = state.conversation_summaries[-1].get("transcript_length", 0)
        return state


class SessionManager:
    """Manages active sessions and their states"""
    def __init__(self):
        self._sessions: dict[str, SessionState] = {}  # session_id -> SessionState
        self._user_active_session: dict[str, str] = {}  # user_id -> active session_id
        self._session_clients: dict[str, set[WebSocket]] = {}  # session_id -> connected clients
        self._client_sessions: dict[WebSocket, str] = {}  # client -> session_id
        self.connected_clients: set[WebSocket] = set()  # All connected clients (for backward compat)

    def get_session_state(self, session_id: str) -> Optional[SessionState]:
        """Get state for a session"""
        return self._sessions.get(session_id)

    def get_or_create_session_state(self, session_id: str, user_id: str) -> SessionState:
        """Get existing or create new session state"""
        if session_id not in self._sessions:
            self._sessions[session_id] = SessionState(session_id, user_id)
        return self._sessions[session_id]

    async def load_session(self, session_id: str, user_id: str) -> Optional[SessionState]:
        """Load session from database into memory, but don't overwrite existing in-memory state"""
        # If already in memory, return the existing state (preserves live data)
        if session_id in self._sessions:
            logger.info(f"Session {session_id} already in memory, using existing state")
            return self._sessions[session_id]

        # Load from database
        session_data = await get_session(session_id, user_id)
        if not session_data:
            return None

        state = SessionState.from_session_data(
            session_id,
            user_id,
            {
                "transcript_entries": session_data.transcript_entries,
                "full_transcript": session_data.full_transcript,
                "detected_questions": session_data.detected_questions,
                "answers": session_data.answers
            }
        )
        self._sessions[session_id] = state
        logger.info(f"Loaded session {session_id} from database: {len(state.transcript)} transcript entries, {len(state.detected_questions)} questions")
        return state

    async def save_session(self, session_id: str) -> bool:
        """Save session state to database"""
        state = self._sessions.get(session_id)
        if not state or not state.dirty:
            return False

        result = await update_session(
            session_id=session_id,
            user_id=state.user_id,
            transcript_entries=state.transcript,
            detected_questions=state.detected_questions,
            answers=state.answers,
            full_transcript=state.full_transcript
        )
        if result:
            state.dirty = False
            state.last_save_time = asyncio.get_event_loop().time()
            return True
        return False

    def set_user_active_session(self, user_id: str, session_id: str):
        """Set the active session for a user"""
        self._user_active_session[user_id] = session_id

    def get_user_active_session(self, user_id: str) -> Optional[str]:
        """Get the active session ID for a user"""
        return self._user_active_session.get(user_id)

    def add_client_to_session(self, client: WebSocket, session_id: str):
        """Associate a client with a session"""
        if session_id not in self._session_clients:
            self._session_clients[session_id] = set()
        self._session_clients[session_id].add(client)
        self._client_sessions[client] = session_id
        self.connected_clients.add(client)
        logger.info(f"[SessionManager] Added client to session {session_id}, total clients for session: {len(self._session_clients[session_id])}")

    def remove_client(self, client: WebSocket):
        """Remove client from session tracking"""
        session_id = self._client_sessions.pop(client, None)
        if session_id and session_id in self._session_clients:
            self._session_clients[session_id].discard(client)
            logger.info(f"[SessionManager] Removed client from session {session_id}, remaining clients: {len(self._session_clients[session_id])}")
        self.connected_clients.discard(client)

    def get_session_clients(self, session_id: str) -> set[WebSocket]:
        """Get all clients connected to a session"""
        return self._session_clients.get(session_id, set())

    def get_client_session(self, client: WebSocket) -> Optional[str]:
        """Get session ID for a client"""
        return self._client_sessions.get(client)

    def unload_session(self, session_id: str):
        """Unload session from memory (after saving)"""
        self._sessions.pop(session_id, None)

    async def auto_save_dirty_sessions(self):
        """Save all dirty sessions (call periodically)"""
        for session_id, state in list(self._sessions.items()):
            if state.dirty:
                await self.save_session(session_id)


# Global session manager
session_manager = SessionManager()

# Global periodic tasks
async def global_periodic_question_detection():
    """Check for questions every 2 seconds using GPT for all active sessions"""
    logger.info("✓ Global periodic question detection started")
    while True:
        await asyncio.sleep(2)
        # Process each active session independently
        for session_id, state in list(session_manager._sessions.items()):
            # Skip temporary sessions (they start with "temp_")
            if not session_id.startswith("temp_"):
                await detect_questions_for_session(session_id, state)

async def global_periodic_session_autosave():
    """Auto-save dirty sessions every 10 seconds"""
    logger.info("✓ Session auto-save started")
    while True:
        await asyncio.sleep(10)
        await session_manager.auto_save_dirty_sessions()


async def summarize_conversation(transcript_text: str) -> str:
    """Generate a summary of the conversation using OpenAI"""
    try:
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        response = client.chat.completions.create(
            model="gpt-4.1-mini",
            max_tokens=500,
            messages=[
                {
                    "role": "system",
                    "content": """You are a conversation summarizer. Create a concise summary of the conversation in bullet points.

INSTRUCTIONS:
- Extract key topics discussed
- Note any important questions asked and answers given
- Highlight any decisions made or action items mentioned
- Keep each bullet point brief (1-2 sentences max)
- Use 3-6 bullet points depending on content length
- Focus on factual information, not filler words

FORMAT:
• [Key point 1]
• [Key point 2]
• [Key point 3]
..."""
                },
                {
                    "role": "user",
                    "content": f"Summarize this conversation:\n\n{transcript_text}"
                }
            ]
        )

        summary = response.choices[0].message.content.strip()
        return summary
    except Exception as e:
        logger.error(f"Failed to generate conversation summary: {e}")
        return ""


async def global_periodic_conversation_summarization():
    """Summarize conversations every 5 minutes for active sessions with new content"""
    logger.info("✓ Conversation summarization started (every 5 minutes)")
    while True:
        await asyncio.sleep(300)  # 5 minutes

        for session_id, state in list(session_manager._sessions.items()):
            # Skip temporary sessions
            if session_id.startswith("temp_"):
                continue

            # Check if there's enough new content to summarize (at least 2000 chars)
            if not state.needs_summary(min_new_chars=2000):
                continue

            try:
                # Get transcript content since last summary
                start_pos = state.last_summary_transcript_length
                new_transcript = state.full_transcript[start_pos:]

                if len(new_transcript.strip()) < 500:  # Skip if too short
                    continue

                logger.info(f"Generating summary for session {session_id} ({len(new_transcript)} new chars)")

                # Generate summary
                summary = await summarize_conversation(new_transcript)

                if summary:
                    state.add_summary(summary)
                    # Save session after adding summary
                    await session_manager.save_session(session_id)

                    # Broadcast summary to clients
                    await broadcast({
                        "type": "conversation_summary",
                        "data": {
                            "summary": summary,
                            "timestamp": datetime.now().isoformat(),
                            "summary_count": len(state.conversation_summaries)
                        }
                    }, session_id)

            except Exception as e:
                logger.error(f"Error summarizing session {session_id}: {e}")


def warmup_services_sync():
    """Warm up external service connections to avoid cold start latency on first query"""
    import time

    # Warm up OpenAI embedding endpoint
    try:
        openai_start = time.time()
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        client.embeddings.create(
            model="text-embedding-3-small",
            input="warmup"
        )
        logger.info(f"✓ OpenAI embeddings warmed up in {time.time() - openai_start:.2f}s")
    except Exception as e:
        logger.warning(f"OpenAI warmup failed (non-critical): {e}")

    # Warm up Pinecone connection
    try:
        pinecone_start = time.time()
        store = PineconeVectorStore()
        if store.is_configured():
            index = store._get_index()
            index.describe_index_stats()
            logger.info(f"✓ Pinecone warmed up in {time.time() - pinecone_start:.2f}s")
        else:
            logger.info("Pinecone not configured, skipping warmup")
    except Exception as e:
        logger.warning(f"Pinecone warmup failed (non-critical): {e}")


async def background_warmup():
    """Run warmup in background so server can accept connections immediately"""
    await asyncio.sleep(0.1)  # Let the server start first
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, warmup_services_sync)

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Initialize database
    logger.info("Initializing database...")
    await db.init_db()
    logger.info("✓ Database initialized")

    # Start periodic tasks when app starts
    logger.info("Starting global periodic tasks...")
    question_task = asyncio.create_task(global_periodic_question_detection())
    autosave_task = asyncio.create_task(global_periodic_session_autosave())
    summarization_task = asyncio.create_task(global_periodic_conversation_summarization())
    logger.info("✓ Global periodic tasks running")

    # Warm up external services in background (non-blocking)
    logger.info("Starting background warmup of external services...")
    warmup_task = asyncio.create_task(background_warmup())

    yield
    # Save all sessions before shutdown
    await session_manager.auto_save_dirty_sessions()
    # Cancel tasks on shutdown
    question_task.cancel()
    autosave_task.cancel()
    summarization_task.cancel()
    warmup_task.cancel()
    # Close database
    await db.close()

app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Health check endpoint for Railway
@app.get("/health")
async def health_check():
    """Health check endpoint - returns 200 when server is ready"""
    return {"status": "healthy"}

# Authentication endpoints
@app.get("/api/auth/config")
async def get_auth_config():
    """Return Clerk configuration for frontend"""
    publishable_key = os.getenv("CLERK_PUBLISHABLE_KEY", "")
    return {
        "publishableKey": publishable_key,
        "isConfigured": clerk_auth.is_configured()
    }

@app.get("/api/auth/me")
async def get_current_user_info(request: Request):
    """Get current authenticated user info"""
    if not clerk_auth.is_configured():
        # Return a mock user when Clerk is not configured (dev mode)
        return {
            "authenticated": True,
            "user": {
                "id": "dev_user",
                "email": "dev@localhost",
                "fullName": "Development User"
            },
            "devMode": True
        }

    user = await optional_auth(request)
    if user:
        return {
            "authenticated": True,
            "user": {
                "id": user.user_id,
                "email": user.email,
                "fullName": user.full_name,
                "firstName": user.first_name,
                "lastName": user.last_name,
                "imageUrl": user.image_url
            },
            "devMode": False
        }
    return {"authenticated": False, "user": None, "devMode": False}


# Cache for users we've already persisted to database (avoids DB call on every request)
_persisted_users: set = set()

async def get_user_id_from_request(request: Request) -> str:
    """Get user ID from request, returns 'dev_user' if auth not configured, raises error if auth fails"""
    if not clerk_auth.is_configured():
        return "dev_user"

    user = await optional_auth(request)
    if user:
        # Only persist to database if we haven't seen this user yet in this session
        if user.user_id not in _persisted_users:
            await get_or_create_user(
                user_id=user.user_id,
                email=user.email,
                first_name=user.first_name,
                last_name=user.last_name,
                image_url=user.image_url
            )
            _persisted_users.add(user.user_id)
        return user.user_id
    # No fallback to anonymous - raise error
    raise HTTPException(status_code=401, detail="Authentication required")


async def get_user_id_from_websocket(websocket: WebSocket, token: str = None) -> str:
    """Get user ID from WebSocket, returns 'dev_user' if auth not configured, raises error if auth fails"""
    if not clerk_auth.is_configured():
        return "dev_user"

    user = await clerk_auth.verify_websocket(websocket, token)
    if user:
        # Only persist to database if we haven't seen this user yet
        if user.user_id not in _persisted_users:
            await get_or_create_user(
                user_id=user.user_id,
                email=user.email,
                first_name=user.first_name,
                last_name=user.last_name,
                image_url=user.image_url
            )
            _persisted_users.add(user.user_id)
        return user.user_id
    # No fallback to anonymous - raise error
    raise HTTPException(status_code=401, detail="WebSocket authentication required")

# Broadcast to all connected frontend clients
async def broadcast(message: dict, session_id: str = None):
    """Broadcast message to clients. If session_id provided, only send to that session's clients."""
    disconnected = set()

    if session_id:
        # Broadcast to session-specific clients
        clients = session_manager.get_session_clients(session_id)
        logger.info(f"Broadcasting to session {session_id}: {len(clients)} clients, msg_type={message.get('type')}")
    else:
        # Broadcast to all connected clients (backward compatibility)
        clients = session_manager.connected_clients
        logger.info(f"Broadcasting to all: {len(clients)} clients, msg_type={message.get('type')}")

    for client in clients:
        try:
            await client.send_json(message)
            logger.debug(f"Sent message to client successfully")
        except Exception as e:
            logger.warning(f"Failed to send message to client: {e}")
            disconnected.add(client)

    for client in disconnected:
        session_manager.remove_client(client)

# Force question detection on selected transcript
async def force_detect_questions(text: str, session_id: str = None, user_id: str = None):
    """Force question detection on user-selected transcript text - generates questions FROM the text"""
    if len(text) < 10:
        return

    # Get the appropriate state - MUST have session_id for proper isolation
    if not session_id or not user_id:
        logger.warning("force_detect_questions called without session_id/user_id - skipping")
        return

    state = session_manager.get_session_state(session_id)
    if not state:
        # Try to load session from database
        state = await session_manager.load_session(session_id, user_id)
        if not state:
            logger.warning(f"Session {session_id} not found for force_detect_questions")
            return

    try:
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        system = """You are a question detector AND generator for interview preparation. Given a transcript excerpt:

YOUR TASK:
1. DETECT any questions already being asked in the text (direct or implicit)
2. GENERATE additional relevant interview questions based on topics mentioned

DETECTION RULES:
- Direct questions: "What is Docker?", "How does it work?"
- Implicit questions: "Tell me about...", "Explain...", "I want to know about..." → convert to proper questions

GENERATION RULES:
- Generate 1-2 interview questions about the main topics/technologies/projects mentioned
- Questions should be interview-style (e.g., "Can you explain...", "How did you implement...")

CRITICAL FORMATTING:
1. ALL questions MUST end with "?" - this is mandatory
2. Make questions specific to the content in the transcript
3. Combine fragmented questions into complete sentences

EXAMPLES:
- "tell me about your RIA project" → "Can you tell me about your RIA project?"
- Text mentions "Docker and Kubernetes" → Generate: "How did you use Docker and Kubernetes in your project?"
- "I want to understand the data pipeline" → "Can you explain your data pipeline architecture?"

Respond with a JSON object containing an array of questions:
{"questions": ["question 1?", "question 2?"]}"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=500,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": f"Selected transcript:\n{text}"}
            ]
        )

        result_text = response.choices[0].message.content.strip()
        logger.info(f"Force question detection GPT response: {result_text}")

        # Remove markdown code blocks if present
        if result_text.startswith('```'):
            lines = result_text.split('\n')
            result_text = '\n'.join(lines[1:-1]) if len(lines) > 2 else result_text
            result_text = result_text.replace('```json', '').replace('```', '').strip()

        json_text = result_text.split('\n\n')[0].strip()
        result = json.loads(json_text)
        questions = result.get('questions', [])
        logger.info(f"Force detection extracted {len(questions)} questions: {questions}")

        # Add all generated questions
        for question_text in questions:
            if question_text and question_text.strip().endswith('?'):
                # Check if question already exists
                exists = any(
                    existing['text'].lower() == question_text.lower()
                    for existing in state.detected_questions
                )
                if not exists:
                    q_entry = {
                        "text": question_text,
                        "speaker": 0,
                        "timestamp": datetime.now().isoformat(),
                        "source": "manual"
                    }
                    state.detected_questions.insert(0, q_entry)
                    state.dirty = True

                    logger.info(f"✓ Generated question from selection: {question_text}")
                    # Broadcast immediately
                    await broadcast({
                        "type": "question_detected",
                        "data": q_entry
                    }, session_id)

        # Auto-save session if we added questions
        if session_id:
            await session_manager.save_session(session_id)

    except Exception as e:
        logger.error(f"Force question detection error: {e}")

# Detect questions using GPT (smarter detection) - Session-aware version
async def detect_questions_for_session(session_id: str, state: SessionState):
    """Detect questions for a specific session"""
    transcript, has_new = state.get_new_transcript_for_detection(2500)

    # Skip if no new content or too short
    if not has_new:
        logger.debug(f"Skipping question detection for session {session_id}: no new transcript content")
        return

    if len(transcript) < 50:
        logger.debug(f"Skipping question detection for session {session_id}: transcript too short ({len(transcript)} chars)")
        return

    logger.info(f"Running question detection for session {session_id} ({len(transcript)} chars): {transcript[:100]}...")

    try:
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        system = """You are a question detector. Analyze the transcript and identify ALL questions being asked by speakers, including both direct questions and implicit questions (requests for information).

TYPES OF QUESTIONS TO DETECT:
1. Direct questions: "What is Docker?", "How does it work?", "Can you explain?"
2. Implicit questions (requests for information): "I would like to know about...", "Tell me about...", "Explain..."

CRITICAL FORMATTING RULES:
1. ALL questions MUST end with "?" - this is mandatory
2. Convert implicit questions to proper question format
3. Questions must be at least 10 characters long
4. Combine fragmented questions across multiple utterances into ONE complete, grammatically correct question
5. IMPORTANT: Detect ALL questions in the transcript, including follow-up questions

EXAMPLES:
Input: "I would like to know more about the Docker implementation"
Output: "What would you like to know about the Docker implementation?" ✓

Input: "Tell me about Kubernetes"
Output: "Can you tell me about Kubernetes?" ✓

Input: "I want to understand how you implemented DevOps"
Output: "How did you implement DevOps?" ✓

Input: "Explain the architecture"
Output: "Can you explain the architecture?" ✓

Input: "What is the status of the project?"
Output: "What is the status of the project?" ✓

Input: "How did you implement the RIA and data pipeline?"
Output: "How did you implement the RIA and data pipeline?" ✓

COMBINING FRAGMENTED QUESTIONS:
- Split: "So I would like to know more about" + "the Docker and Kubernetes implementation"
  → Combined: "What would you like to know about the Docker and Kubernetes implementation?" ✓

INVALID formats to IGNORE (will NOT be detected):
✗ General statements without information request
✗ Comments, greetings, acknowledgments

Respond with a JSON object containing ALL questions found:
{"questions": ["question 1?", "question 2?", "question 3?"]}

If no questions found, return {"questions": []}"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=500,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": f"Recent transcript:\n{transcript}"}
            ]
        )

        text = response.choices[0].message.content
        if not text:
            logger.warning("GPT returned empty response")
            return

        text = text.strip()
        logger.info(f"Raw GPT response: {text}")

        # Remove markdown code blocks if present
        if text.startswith('```'):
            # Extract JSON from code block
            lines = text.split('\n')
            text = '\n'.join(lines[1:-1]) if len(lines) > 2 else text
            text = text.replace('```json', '').replace('```', '').strip()

        # Extract only the JSON part (before any explanatory text after double newline)
        # Haiku sometimes adds explanations after the JSON
        json_text = text.split('\n\n')[0].strip()

        if not json_text:
            logger.warning("No JSON text found in GPT response")
            return

        result = json.loads(json_text)
        questions = result.get('questions', [])

        # Process all detected questions
        for question_text in questions:
            if question_text:
                # Validate that it's actually a question using our regex
                if not detect_question(question_text):
                    logger.info(f"GPT detected non-question format, skipping: {question_text}")
                else:
                    # Check if question already exists IN THIS SESSION'S questions
                    exists = any(
                        existing['text'].lower() == question_text.lower()
                        for existing in state.detected_questions
                    )
                    if not exists:
                        q_entry = {
                            "text": question_text,
                            "speaker": 0,
                            "timestamp": datetime.now().isoformat(),
                            "source": "gpt"
                        }
                        state.detected_questions.insert(0, q_entry)
                        state.detected_questions = state.detected_questions[:10]  # Keep more per session
                        state.dirty = True

                        logger.info(f"✓ GPT detected question for session {session_id}: {question_text}")
                        # Broadcast ONLY to this session's clients
                        await broadcast({
                            "type": "question_detected",
                            "data": q_entry
                        }, session_id)

        # Save session after adding questions
        if state.dirty:
            await session_manager.save_session(session_id)

        # Log usage for question detection
        try:
            await log_usage(
                user_id=state.user_id,
                service="openai",
                operation="question_detection",
                tokens=len(transcript.split())  # Approximate input tokens
            )
        except Exception as log_err:
            logger.warning(f"Failed to log question detection usage: {log_err}")

    except Exception as e:
        logger.error(f"Question detection error for session {session_id}: {e}")
        if 'text' in locals():
            logger.error(f"Failed to parse GPT response: {text[:200]}")

# Note: analyze_transcript() removed - was using global state
# Suggestions feature is disabled; if re-enabled, make it session-aware

# Answer question with GPT (streaming with proper event handling)
async def answer_question_stream(question: str, websocket: WebSocket, user_id: str, session_id: str):
    """Answer a question with GPT streaming. Requires user_id and session_id."""
    import time
    timings = {}
    total_start = time.time()

    if not session_id:
        raise ValueError("session_id is required for answer_question_stream")

    full_answer = ""

    # Get session state - required
    session_state = session_manager.get_session_state(session_id)
    if not session_state:
        raise ValueError(f"Session state not found for session_id={session_id}")

    try:
        # === STEP 1: Pinecone Search (Retrieval) ===
        retrieval_start = time.time()

        store_start = time.time()
        store = get_vector_store_for_user(user_id)
        logger.info(f"[TIMING] get_vector_store_for_user: {time.time() - store_start:.2f}s")

        logger.info(f"[TIMING] Starting Pinecone search for: {question[:50]}...")
        search_start = time.time()
        search_results = await store.search(user_id, question, top_k=8)
        logger.info(f"[TIMING] store.search() returned: {time.time() - search_start:.2f}s")

        timings['retrieval'] = time.time() - retrieval_start
        logger.info(f"[TIMING] Pinecone retrieval (total): {timings['retrieval']:.2f}s")

        # Filter results by score threshold for better relevance
        relevant_results = [r for r in search_results if r.score > 0.3]
        logger.info(f"Found {len(search_results)} results, {len(relevant_results)} above threshold")

        # Log the results for debugging
        for i, r in enumerate(relevant_results):
            logger.info(f"  Result {i+1}: doc='{r.document_name}', score={r.score:.3f}, content={r.content[:100]}...")

        doc_context = store.get_context(relevant_results)

        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        doc_section = f"""PROJECT DOCUMENTS (Use these to answer questions about the user's projects):
{doc_context}""" if doc_context else "No project documents uploaded yet."

        system = f"""You are an expert interview coach helping an Indian IT professional answer technical interview questions in real-time. Your responses should sound natural and conversational, like how an experienced Indian software engineer would explain things.

SPEAKING STYLE:
- Use natural, conversational English as spoken by Indian IT professionals
- Phrases like "So basically...", "What we did is...", "The thing is...", "Actually..." are fine
- Be confident and direct, like explaining to a colleague
- Use "we" when talking about project work (e.g., "We implemented...", "What we used is...")
- Avoid overly formal or robotic language

RESPONSE GUIDELINES:
1. **Start directly with the answer** - no preamble like "Great question" or "Sure"
2. **Be comprehensive but concise** - provide enough detail to demonstrate expertise (4-6 sentences for simple topics, more for complex ones)
3. **Use the document context** - when answering about projects, USE SPECIFIC DETAILS from the documents provided
4. **Structure for clarity** - use bullet points for multiple components
5. **Include concrete examples** - relate to the actual project when documents are available

WHEN DOCUMENTS ARE PROVIDED:
- Extract and cite SPECIFIC technical details (technologies, architectures, frameworks)
- Mention actual features, components, or implementations from the documents
- Reference specific aspects like data pipelines, APIs, databases, or workflows mentioned
- Provide depth by explaining HOW things work based on document details

RESPONSE FORMAT:
- For project questions: Start with a brief overview, then explain key technical components with specifics from documents
- For concept questions: Define clearly, explain how it works, give a practical example
- Use bullet points for listing multiple features or components

{doc_section}"""

        # Get transcript from session (session_state is guaranteed to exist now)
        transcript = session_state.get_recent_transcript(2000)

        # Get conversation summaries for broader context
        summaries_text = session_state.get_summaries_text()
        summaries_count = len(session_state.conversation_summaries)
        logger.info(f"Including {summaries_count} conversation summaries in context")

        # Build conversation context with summaries + recent transcript
        conversation_context = ""
        if summaries_text:
            conversation_context += f"{summaries_text}\n\n"
        conversation_context += f"RECENT CONVERSATION:\n{transcript}"

        # Send start event
        await websocket.send_json({
            "type": "answer_start",
            "data": {"question": question}
        })

        # === STEP 2: OpenAI API Call (Generation) ===
        generation_start = time.time()
        logger.info(f"[TIMING] Starting OpenAI stream...")
        chunk_count = 0
        first_chunk_time = None

        # Create stream and iterate chunks (OpenAI streaming)
        stream = client.chat.completions.create(
            model="gpt-4.1-mini",
            max_tokens=2000,
            stream=True,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": f"Conversation context:\n{conversation_context}\n\nQuestion to answer: {question}\n\nProvide a detailed, expert-level response using any relevant document context. Use the conversation summaries to understand the broader context of the discussion."}
            ]
        )

        finish_reason = None
        for chunk in stream:
            if first_chunk_time is None and chunk.choices[0].delta.content:
                first_chunk_time = time.time()
                timings['time_to_first_chunk'] = first_chunk_time - generation_start
                logger.info(f"[TIMING] Time to first chunk: {timings['time_to_first_chunk']:.2f}s")
            if chunk.choices[0].delta.content:
                text = chunk.choices[0].delta.content
                chunk_count += 1
                full_answer += text
                logger.debug(f"Chunk {chunk_count}: {len(text)} chars")

                # Send chunk to client
                try:
                    await websocket.send_json({
                        "type": "answer_chunk",
                        "data": {
                            "question": question,
                            "chunk": text,
                            "is_final": False
                        }
                    })
                    # Small delay to ensure chunks are sent separately
                    await asyncio.sleep(0.01)
                except Exception as e:
                    logger.error(f"Error sending chunk: {e}")

            if chunk.choices[0].finish_reason:
                finish_reason = chunk.choices[0].finish_reason

        # Calculate generation timing
        timings['total_generation'] = time.time() - generation_start
        logger.info(f"[TIMING] Total generation: {timings['total_generation']:.2f}s ({chunk_count} chunks)")

        # Send final message with complete answer
        await websocket.send_json({
            "type": "answer_chunk",
            "data": {
                "question": question,
                "chunk": "",
                "is_final": True,
                "full_answer": full_answer,
                "stop_reason": finish_reason
            }
        })

        # Save answer to session if available
        if session_state and full_answer:
            session_state.add_answer(question, full_answer)
            await session_manager.save_session(session_id)

            # Log usage for answer generation
            try:
                await log_usage(
                    user_id=user_id,
                    service="openai",
                    operation="completion",
                    tokens=len(full_answer.split())  # Approximate token count
                )
            except Exception as log_err:
                logger.warning(f"Failed to log usage: {log_err}")

        # === TIMING SUMMARY ===
        timings['total'] = time.time() - total_start
        logger.info(f"[TIMING] === ANSWER GENERATION SUMMARY ===")
        logger.info(f"[TIMING]   Pinecone retrieval: {timings.get('retrieval', 0):.2f}s")
        logger.info(f"[TIMING]   Time to first chunk: {timings.get('time_to_first_chunk', 0):.2f}s")
        logger.info(f"[TIMING]   Total generation: {timings.get('total_generation', 0):.2f}s")
        logger.info(f"[TIMING]   Total time: {timings['total']:.2f}s")
        logger.info(f"[TIMING] ================================")

    except Exception as e:
        logger.error(f"Answer error: {e}")
        import traceback
        traceback.print_exc()

        # Send error with partial answer if available
        error_msg = f"Error occurred: {str(e)}"
        if full_answer:
            error_msg = f"{full_answer}\n\n[Error: Response interrupted - {str(e)}]"

        await websocket.send_json({
            "type": "answer_chunk",
            "data": {
                "question": question,
                "chunk": "",
                "is_final": True,
                "full_answer": error_msg,
                "error": str(e)
            }
        })

# WebSocket for frontend clients (receive updates)
@app.websocket("/ws/client")
async def client_websocket(websocket: WebSocket, token: str = Query(None), session_id: str = Query(None)):
    await websocket.accept()

    # Get user ID for this connection
    user_id = await get_user_id_from_websocket(websocket, token)
    logger.info(f"Client WebSocket connected: user_id={user_id}, session_id={session_id}")

    # Session ID is required - client should create session via API first
    if not session_id:
        await websocket.send_json({"type": "error", "message": "session_id is required. Create a session first via POST /api/sessions"})
        await websocket.close(code=4000, reason="session_id required")
        return

    current_session_id = session_id

    # Try to load existing session
    session_state = await session_manager.load_session(session_id, user_id)
    if not session_state:
        # Create session state for this session (might be newly created)
        session_state = session_manager.get_or_create_session_state(session_id, user_id)
        logger.info(f"Created new session state for session_id={session_id}")

    # Register client with session manager
    session_manager.add_client_to_session(websocket, current_session_id)
    session_manager.set_user_active_session(user_id, current_session_id)

    # Send session data to client
    init_data = {
        "transcript": session_state.transcript,
        "questions": session_state.detected_questions,
        "suggestions": session_state.suggestions,
        "userId": user_id,
        "sessionId": current_session_id
    }

    await websocket.send_json({"type": "init", "data": init_data})

    try:
        while True:
            data = await websocket.receive_json()

            if data.get("type") == "get_answer":
                question = data.get("question", "")
                await answer_question_stream(question, websocket, user_id, current_session_id)

            elif data.get("type") == "force_question_detection":
                text = data.get("text", "")
                logger.info(f"Received force_question_detection request: text_length={len(text) if text else 0}")
                if text:
                    await force_detect_questions(text, current_session_id, user_id)
                else:
                    logger.warning("force_question_detection received with empty text")

            elif data.get("type") == "clear":
                # Clear session state (session is always available now)
                session_state.clear()
                await session_manager.save_session(current_session_id)
                await broadcast({"type": "cleared"}, current_session_id)

            elif data.get("type") == "switch_session":
                # Client is switching to a different session
                new_session_id = data.get("session_id")
                if new_session_id and new_session_id != current_session_id:
                    logger.info(f"Switching session from {current_session_id} to {new_session_id}")

                    # Save current session before switching
                    if current_session_id:
                        await session_manager.save_session(current_session_id)
                        session_manager.remove_client(websocket)

                    # Load new session
                    current_session_id = new_session_id
                    session_state = await session_manager.load_session(new_session_id, user_id)

                    if session_state:
                        session_manager.add_client_to_session(websocket, new_session_id)
                        session_manager.set_user_active_session(user_id, new_session_id)

                        # Send new session data
                        await websocket.send_json({
                            "type": "session_loaded",
                            "data": {
                                "transcript": session_state.transcript,
                                "questions": session_state.detected_questions,
                                "suggestions": session_state.suggestions,
                                "sessionId": new_session_id
                            }
                        })
                        logger.info(f"Session {new_session_id} loaded: {len(session_state.transcript)} transcript entries, {len(session_state.detected_questions)} questions")
                    else:
                        # Session doesn't exist in database
                        logger.warning(f"Session {new_session_id} not found in database")
                        await websocket.send_json({
                            "type": "session_not_found",
                            "data": {
                                "sessionId": new_session_id,
                                "message": "Session not found"
                            }
                        })

            elif data.get("type") == "create_session":
                # Create a new session
                title = data.get("title")
                new_session = await create_session(user_id, title)

                # Save and switch from current session
                if current_session_id:
                    await session_manager.save_session(current_session_id)
                    session_manager.remove_client(websocket)

                current_session_id = new_session.id
                session_state = session_manager.get_or_create_session_state(new_session.id, user_id)
                session_manager.add_client_to_session(websocket, new_session.id)
                session_manager.set_user_active_session(user_id, new_session.id)

                await websocket.send_json({
                    "type": "session_created",
                    "data": {
                        "id": new_session.id,
                        "title": new_session.title,
                        "sessionId": new_session.id
                    }
                })

    except WebSocketDisconnect:
        # Save session on disconnect
        if current_session_id:
            await session_manager.save_session(current_session_id)
        session_manager.remove_client(websocket)
        logger.info(f"Client WebSocket disconnected: user_id={user_id}, session_id={current_session_id}")

# WebSocket for audio streaming with Deepgram
@app.websocket("/ws/audio")
async def audio_websocket(websocket: WebSocket, token: str = Query(None), session_id: str = Query(None)):
    await websocket.accept()

    # Session ID is required - client should create session via API first
    if not session_id:
        await websocket.send_json({"type": "error", "message": "session_id is required. Create a session first via POST /api/sessions"})
        await websocket.close(code=4000, reason="session_id required")
        return

    # Send immediate acknowledgment so client knows connection is alive
    await websocket.send_json({"type": "connecting", "message": "Connecting to transcription service..."})

    # Get user ID for this connection
    user_id = await get_user_id_from_websocket(websocket, token)
    logger.info(f"Audio WebSocket connected: user_id={user_id}, session_id={session_id}")

    # Get or create session state
    session_state = session_manager.get_session_state(session_id)
    if not session_state:
        session_state = await session_manager.load_session(session_id, user_id)
    if not session_state:
        session_state = session_manager.get_or_create_session_state(session_id, user_id)
        logger.info(f"Created new session state for audio: session_id={session_id}")

    # Initialize Deepgram client (SDK v5)
    deepgram = DeepgramClient(api_key=os.getenv("DEEPGRAM_API_KEY"))

    # Capture the event loop from the async context
    loop = asyncio.get_running_loop()

    # Track current utterance for interim results
    last_processed = ""

    try:
        # Use context manager for SDK v5 connection
        logger.info("Connecting to Deepgram v5...")

        with deepgram.listen.v1.connect(
            model="nova-2",
            language="en",
            encoding="linear16",
            sample_rate="48000",  # Browser native sample rate (usually 48kHz)
            channels="1",
            smart_format="true",
            punctuate="true",
            interim_results="true",
            utterance_end_ms="1000",
            vad_events="true",
        ) as dg_connection:

            # Define event handlers for SDK v5
            async def on_message_async(message):
                nonlocal last_processed
                try:
                    transcript = ""
                    is_final = False

                    # SDK v5 returns message with channel.alternatives
                    if hasattr(message, 'channel') and message.channel:
                        if hasattr(message.channel, 'alternatives') and message.channel.alternatives:
                            transcript = message.channel.alternatives[0].transcript or ""

                    is_final = getattr(message, 'is_final', False) or getattr(message, 'speech_final', False)

                    logger.info(f"Deepgram transcript: '{transcript[:50] if transcript else ''}', is_final={is_final}")

                    if not transcript:
                        return

                    # No speaker diarization - use single speaker
                    speaker = 0

                    # Broadcast to the same session - "temp" session will broadcast to clients in "temp" session
                    broadcast_session = session_id

                    if is_final:
                        # Final transcript - always use session state
                        session_state.add_utterance(speaker, transcript, is_final=True)

                        # Check for questions in real-time (keyword-based)
                        question = detect_question(transcript)
                        if question and question != last_processed:
                            last_processed = question
                            q_entry = {
                                "text": question,
                                "speaker": speaker,
                                "timestamp": datetime.now().isoformat(),
                                "source": "keyword"
                            }
                            session_state.detected_questions.insert(0, q_entry)
                            session_state.detected_questions = session_state.detected_questions[:10]
                            session_state.dirty = True

                            # Broadcast question to clients
                            await broadcast({
                                "type": "question_detected",
                                "data": q_entry
                            }, broadcast_session)

                        # Broadcast transcript update
                        await broadcast({
                            "type": "transcript",
                            "data": {
                                "speaker": speaker,
                                "text": transcript,
                                "is_final": True
                            }
                        }, broadcast_session)
                    else:
                        # Interim result
                        await broadcast({
                            "type": "transcript",
                            "data": {
                                "speaker": speaker,
                                "text": transcript,
                                "is_final": False
                            }
                        }, broadcast_session)

                except Exception as e:
                    logger.error(f"Transcript processing error: {e}")
                    import traceback
                    traceback.print_exc()

            # Sync wrapper for async handler
            def on_message(message):
                asyncio.run_coroutine_threadsafe(on_message_async(message), loop)

            # Register event handlers (SDK v5 uses EventType)
            dg_connection.on(EventType.OPEN, lambda _: logger.info("Deepgram connection opened"))
            dg_connection.on(EventType.MESSAGE, on_message)
            dg_connection.on(EventType.ERROR, lambda error: logger.error(f"Deepgram error: {error}"))
            dg_connection.on(EventType.CLOSE, lambda _: logger.info("Deepgram connection closed"))

            # Start listening in a separate thread
            import threading
            def listening_thread():
                try:
                    dg_connection.start_listening()
                except Exception as e:
                    logger.error(f"Error in listening thread: {e}")

            listen_thread = threading.Thread(target=listening_thread)
            listen_thread.start()

            logger.info("✓ Deepgram v5 connected")

            await websocket.send_json({"type": "ready"})
            logger.info("Sent 'ready' to client, waiting for audio...")

            # Receive and forward audio data to Deepgram
            audio_packet_count = 0
            while True:
                data = await websocket.receive_bytes()
                audio_packet_count += 1
                # Log first few packets and then periodically
                if audio_packet_count <= 3 or audio_packet_count % 100 == 0:
                    # Check if audio has actual content (not silence)
                    import struct
                    if len(data) >= 2:
                        # Sample a few values from the 16-bit PCM data
                        samples = struct.unpack(f'<{min(10, len(data)//2)}h', data[:min(20, len(data))])
                        max_sample = max(abs(s) for s in samples) if samples else 0
                        logger.info(f"Audio packet #{audio_packet_count}: {len(data)} bytes, max_sample={max_sample}, first_samples={samples[:5]}")
                # SDK v5 requires wrapping audio in ListenV1MediaMessage
                dg_connection.send_media(ListenV1MediaMessage(data))

    except WebSocketDisconnect:
        logger.info("Audio WebSocket disconnected")
    except Exception as e:
        logger.error(f"Audio WebSocket error: {e}", exc_info=True)
        try:
            await websocket.send_json({"type": "error", "message": str(e)})
        except Exception as send_err:
            logger.debug(f"Could not send error to client (already disconnected?): {send_err}")

# Document endpoints
@app.get("/api/documents")
async def list_documents(request: Request):
    user_id = await get_user_id_from_request(request)
    # Use PostgreSQL for fast document listing instead of Pinecone
    from database import get_user_documents
    documents = await get_user_documents(user_id)
    return {"documents": documents, "userId": user_id}

@app.post("/api/documents")
async def upload_document(
    request: Request,
    file: Optional[UploadFile] = File(None),
    content: Optional[str] = Form(None),
    name: Optional[str] = Form(None),
    session_id: Optional[str] = Form(None)
):
    user_id = await get_user_id_from_request(request)
    store = get_vector_store_for_user(user_id)

    # session_id is required
    if not session_id:
        raise HTTPException(400, "session_id is required for document upload")

    if file:
        # Use filename if name not provided
        doc_name = name or file.filename
        file_bytes = await file.read()

        # Extract text from file
        try:
            text_content = extract_text_from_file(file_bytes, doc_name)
        except Exception as e:
            raise HTTPException(400, f"Failed to extract text from file: {str(e)}")

        # Get file type from extension
        file_type = doc_name.lower().split('.')[-1] if '.' in doc_name else 'unknown'

        # Add to Pinecone and PostgreSQL
        doc = await store.add_document(
            user_id=user_id,
            name=doc_name,
            content=text_content,
            session_id=session_id,
            file_name=doc_name,
            file_type=file_type
        )
    elif content:
        if not name:
            raise HTTPException(400, "Name required when uploading text content")
        doc = await store.add_document(
            user_id=user_id,
            name=name,
            content=content,
            session_id=session_id,
            file_type='txt'
        )
    else:
        raise HTTPException(400, "No content provided")

    return {
        "id": doc.id,
        "name": doc.name,
        "uploaded_at": doc.uploaded_at.isoformat(),
        "chunks_count": len(doc.chunks),
        "session_id": session_id,
        "userId": user_id
    }

@app.delete("/api/documents/{doc_id}")
async def delete_document(request: Request, doc_id: str):
    user_id = await get_user_id_from_request(request)
    store = get_vector_store_for_user(user_id)

    if not await store.delete_document(user_id, doc_id):
        raise HTTPException(404, "Document not found")
    return {"success": True}

# ============= Session Endpoints =============
# Note: Legacy /api/state and /api/clear endpoints removed - use session-specific endpoints instead

@app.get("/api/sessions")
async def list_sessions(request: Request, include_archived: bool = False):
    """List all sessions for the current user"""
    user_id = await get_user_id_from_request(request)
    sessions = await get_user_sessions(user_id, include_archived)
    return {"sessions": sessions}


@app.post("/api/sessions")
async def create_new_session(request: Request):
    """Create a new session"""
    user_id = await get_user_id_from_request(request)
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    title = body.get("title")

    session = await create_session(user_id, title)

    # Audit log for session creation
    try:
        await log_audit(
            action="session_create",
            user_id=user_id,
            resource_type="session",
            resource_id=session.id,
            extra_data={"title": session.title}
        )
    except Exception as e:
        logger.warning(f"Failed to log audit: {e}")

    return {
        "id": session.id,
        "title": session.title,
        "status": session.status,
        "created_at": session.created_at.isoformat() if session.created_at else None
    }


@app.get("/api/sessions/{session_id}")
async def get_session_details(session_id: str, request: Request):
    """Get full session data including transcript, questions, and answers"""
    user_id = await get_user_id_from_request(request)
    session = await get_session(session_id, user_id)

    if not session:
        raise HTTPException(404, "Session not found")

    return {
        "id": session.id,
        "title": session.title,
        "status": session.status,
        "transcript_entries": session.transcript_entries or [],
        "detected_questions": session.detected_questions or [],
        "answers": session.answers or [],
        "document_ids": session.document_ids or [],
        "full_transcript": session.full_transcript or "",
        "created_at": session.created_at.isoformat() if session.created_at else None,
        "updated_at": session.updated_at.isoformat() if session.updated_at else None
    }


@app.put("/api/sessions/{session_id}")
async def update_session_endpoint(session_id: str, request: Request):
    """Update session data (title, status, etc.)"""
    user_id = await get_user_id_from_request(request)
    body = await request.json()

    session = await update_session(
        session_id=session_id,
        user_id=user_id,
        title=body.get("title"),
        status=body.get("status"),
        transcript_entries=body.get("transcript_entries"),
        detected_questions=body.get("detected_questions"),
        answers=body.get("answers"),
        document_ids=body.get("document_ids"),
        full_transcript=body.get("full_transcript")
    )

    if not session:
        raise HTTPException(404, "Session not found")

    return {"success": True, "updated_at": session.updated_at.isoformat()}


@app.delete("/api/sessions/{session_id}")
async def delete_session_endpoint(session_id: str, request: Request):
    """Delete a session"""
    user_id = await get_user_id_from_request(request)

    if not await delete_session(session_id, user_id):
        raise HTTPException(404, "Session not found")

    # Also unload from memory to prevent stale data
    session_manager.unload_session(session_id)

    # Audit log for session deletion
    try:
        await log_audit(
            action="session_delete",
            user_id=user_id,
            resource_type="session",
            resource_id=session_id
        )
    except Exception as e:
        logger.warning(f"Failed to log audit: {e}")

    return {"success": True}


@app.post("/api/sessions/{session_id}/archive")
async def archive_session_endpoint(session_id: str, request: Request):
    """Archive a session instead of deleting"""
    user_id = await get_user_id_from_request(request)

    if not await archive_session(session_id, user_id):
        raise HTTPException(404, "Session not found")

    # Also unload from memory to prevent stale data
    session_manager.unload_session(session_id)

    # Audit log for session archive
    try:
        await log_audit(
            action="session_archive",
            user_id=user_id,
            resource_type="session",
            resource_id=session_id
        )
    except Exception as e:
        logger.warning(f"Failed to log audit: {e}")

    return {"success": True}


# Note: Answers are now streamed via WebSocket only (see /ws/client endpoint)

# Serve frontend
frontend_dir = os.path.join(os.path.dirname(__file__), "frontend")
if os.path.exists(frontend_dir):
    app.mount("/static", StaticFiles(directory=frontend_dir), name="static")

    @app.get("/")
    async def serve_frontend():
        return FileResponse(
            os.path.join(frontend_dir, "index.html"),
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
